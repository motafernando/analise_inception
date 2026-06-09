import csv
import json
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl


BASE = Path(__file__).resolve().parents[1]
ANALYZER = BASE / "analyzer"
ZIP_REF = ANALYZER / "ultimatum-ner-dataset2390907466170657522.zip"
ZIP_CANDIDATES = ANALYZER / "selecao-ceia-ultimatum4959528506098853350.zip"
ACCOUNTS_XLSX = BASE / "outputs" / "relacao_contas_horario_textos_ultimatum_ceia.xlsx"
OUT_DIR = BASE / "outputs" / "accuracy_report"
ENTITY_LAYER_INDEX = 3
REFERENCE_USERS = ["jacques", "lauana"]
SYSTEM_USERS = {"INITIAL_CAS", "admin", "usuarioteste"}


ACCOUNT_USER_CORRECTIONS = {
    "costajoyce": "Maria Luiza Moreira Macedo",
}


def normalize_doc_name(value):
    value = (value or "").strip()
    value = value.replace(" teste", "")
    if not value:
        return ""
    if value.startswith("vr-46") and "gold" in value:
        return "vr-46-gold.tsv"
    if not value.endswith(".txt") and not value.endswith(".tsv"):
        value = f"{value}.txt"
    return value


def read_accounts():
    workbook = openpyxl.load_workbook(ACCOUNTS_XLSX, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    indexes = {name: idx for idx, name in enumerate(header) if name}
    accounts = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        candidate = row[indexes["Candidato"]] if "Candidato" in indexes else None
        user = row[indexes["Usuário"]] if "Usuário" in indexes else None
        shift = row[indexes["Horário"]] if "Horário" in indexes else None
        texts = row[indexes["TEXTOS"]] if "TEXTOS" in indexes else None
        if not user:
            continue

        user = str(user).strip()
        candidate = ACCOUNT_USER_CORRECTIONS.get(user, str(candidate).strip() if candidate else user)
        assigned_docs = [normalize_doc_name(part) for part in str(texts or "").split(";")]
        assigned_docs = [doc for doc in assigned_docs if doc]
        accounts[user] = {
            "usuario": user,
            "candidato": candidate,
            "horario": str(shift).strip() if shift else "",
            "textos_planilha": str(texts or "").strip(),
            "documentos_planilha": assigned_docs,
            "corrigido": user in ACCOUNT_USER_CORRECTIONS,
        }

    return accounts


def annotation_files(zip_path):
    docs = defaultdict(list)
    with zipfile.ZipFile(zip_path) as archive:
        for name in archive.namelist():
            parts = name.split("/")
            if len(parts) == 3 and parts[0] == "annotation" and parts[2].endswith(".tsv"):
                docs[parts[1]].append(parts[2][:-4])
    return {doc: sorted(users) for doc, users in sorted(docs.items())}


def read_labels(zip_path, document, user):
    path = f"annotation/{document}/{user}.tsv"
    tokens = []
    labels = []

    with zipfile.ZipFile(zip_path) as archive:
        raw = archive.read(path).decode("utf-8")

    for line in raw.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue

        columns = line.split("\t")
        if len(columns) <= ENTITY_LAYER_INDEX or "-" not in columns[0]:
            continue

        label = columns[ENTITY_LAYER_INDEX]
        if label in {"_", "*"}:
            label = "O"
        elif "[" in label:
            label = label.split("[", 1)[0]

        tokens.append(columns[2])
        labels.append(label)

    return tokens, labels


def read_label_map(zip_path):
    label_map = {"O": "O"}
    pattern = re.compile(r"<([^>]+)>.*?rdfs:label>\s+\"([^\"]+)\"", re.DOTALL)

    with zipfile.ZipFile(zip_path) as archive:
        ttl_files = [name for name in archive.namelist() if name.startswith("kb/") and name.endswith(".ttl")]
        for name in ttl_files:
            text = archive.read(name).decode("utf-8", errors="replace")
            text = text.replace("<http://www.w3.org/2000/01/rdf-schema#label>", "rdfs:label>")
            for uri, label in pattern.findall(text):
                label_map[uri] = label

    return label_map


def label_name(label, label_map):
    return label_map.get(label, label.rsplit("#", 1)[-1] if "#" in label else label)


def compute_accuracy(reference_labels, candidate_labels):
    total = len(reference_labels)
    correct = sum(a == b for a, b in zip(reference_labels, candidate_labels))
    reference_entity_indexes = [
        idx for idx, label in enumerate(reference_labels)
        if label != "O"
    ]
    reference_entity_correct = sum(
        reference_labels[idx] == candidate_labels[idx]
        for idx in reference_entity_indexes
    )
    balanced_indexes = [
        idx for idx, (a, b) in enumerate(zip(reference_labels, candidate_labels))
        if a != "O" or b != "O"
    ]
    balanced_correct = sum(reference_labels[idx] == candidate_labels[idx] for idx in balanced_indexes)
    ignored_both_o = total - len(balanced_indexes)
    return {
        "total_tokens": total,
        "acertos": correct,
        "acuracia": correct / total if total else None,
        "tokens_entidade_referencia": len(reference_entity_indexes),
        "acertos_entidade_referencia": reference_entity_correct,
        "acuracia_entidade_referencia": (
            reference_entity_correct / len(reference_entity_indexes)
            if reference_entity_indexes else None
        ),
        "tokens_balanceados": len(balanced_indexes),
        "tokens_ignorados_o_o": ignored_both_o,
        "acertos_balanceados": balanced_correct,
        "acuracia_balanceada": balanced_correct / len(balanced_indexes) if balanced_indexes else None,
    }


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    accounts = read_accounts()
    ref_docs = annotation_files(ZIP_REF)
    candidate_docs = annotation_files(ZIP_CANDIDATES)
    label_map = read_label_map(ZIP_REF)

    details = []
    skipped = []
    label_counts = []

    for document, users in candidate_docs.items():
        if document not in ref_docs:
            skipped.append({
                "documento": document,
                "usuario": "",
                "motivo": "Documento sem anotacao de referencia no Projeto A",
            })
            continue

        valid_refs = [user for user in REFERENCE_USERS if user in ref_docs[document]]
        if not valid_refs:
            skipped.append({
                "documento": document,
                "usuario": "",
                "motivo": "Documento sem jacques/lauana como referencia",
            })
            continue

        for candidate_user in users:
            if candidate_user in SYSTEM_USERS:
                skipped.append({
                    "documento": document,
                    "usuario": candidate_user,
                    "motivo": "Usuario tecnico/sistema ignorado",
                })
                continue

            try:
                candidate_tokens, candidate_labels = read_labels(ZIP_CANDIDATES, document, candidate_user)
            except KeyError:
                skipped.append({"documento": document, "usuario": candidate_user, "motivo": "TSV nao encontrado"})
                continue

            if not candidate_tokens:
                skipped.append({"documento": document, "usuario": candidate_user, "motivo": "TSV sem tokens anotaveis"})
                continue

            account = accounts.get(candidate_user, {})
            candidate_label_counts = Counter(candidate_labels)
            for raw_label, count in sorted(candidate_label_counts.items(), key=lambda item: (-item[1], item[0])):
                label_counts.append({
                    "documento": document,
                    "usuario": candidate_user,
                    "candidato": account.get("candidato", candidate_user),
                    "rotulo": label_name(raw_label, label_map),
                    "contagem": count,
                })

            for ref_user in valid_refs:
                reference_tokens, reference_labels = read_labels(ZIP_REF, document, ref_user)
                if len(reference_tokens) != len(candidate_tokens):
                    skipped.append({
                        "documento": document,
                        "usuario": candidate_user,
                        "referencia": ref_user,
                        "motivo": f"Tokens divergentes: referencia={len(reference_tokens)} candidato={len(candidate_tokens)}",
                    })
                    continue

                token_mismatches = sum(a != b for a, b in zip(reference_tokens, candidate_tokens))
                if token_mismatches:
                    skipped.append({
                        "documento": document,
                        "usuario": candidate_user,
                        "referencia": ref_user,
                        "motivo": f"Textos de tokens divergentes em {token_mismatches} posicoes",
                    })
                    continue

                metrics = compute_accuracy(reference_labels, candidate_labels)
                details.append({
                    "documento": document,
                    "referencia": ref_user,
                    "usuario": candidate_user,
                    "candidato": account.get("candidato", candidate_user),
                    "horario": account.get("horario", ""),
                    "usuario_corrigido": "sim" if account.get("corrigido") else "nao",
                    "acuracia": metrics["acuracia"],
                    "acuracia_geral_percentual": metrics["acuracia"] * 100 if metrics["acuracia"] is not None else None,
                    "acertos": metrics["acertos"],
                    "total_tokens": metrics["total_tokens"],
                    "acuracia_entidade_referencia": metrics["acuracia_entidade_referencia"],
                    "acuracia_entidade_referencia_percentual": (
                        metrics["acuracia_entidade_referencia"] * 100
                        if metrics["acuracia_entidade_referencia"] is not None else None
                    ),
                    "acertos_entidade_referencia": metrics["acertos_entidade_referencia"],
                    "tokens_entidade_referencia": metrics["tokens_entidade_referencia"],
                    "acuracia_balanceada": metrics["acuracia_balanceada"],
                    "acuracia_balanceada_percentual": (
                        metrics["acuracia_balanceada"] * 100
                        if metrics["acuracia_balanceada"] is not None else None
                    ),
                    "acertos_balanceados": metrics["acertos_balanceados"],
                    "tokens_balanceados": metrics["tokens_balanceados"],
                    "tokens_ignorados_o_o": metrics["tokens_ignorados_o_o"],
                })

    summary_map = {}
    for row in details:
        key = (row["usuario"], row["candidato"], row["horario"], row["documento"])
        item = summary_map.setdefault(key, {
            "usuario": row["usuario"],
            "candidato": row["candidato"],
            "horario": row["horario"],
            "documento": row["documento"],
            "qtd_referencias": 0,
            "media_acuracia_geral": 0.0,
            "media_acuracia_entidade_referencia": 0.0,
            "media_acuracia_balanceada": 0.0,
            "total_tokens_medio": 0.0,
            "tokens_entidade_referencia_medio": 0.0,
            "tokens_balanceados_medio": 0.0,
            "tokens_ignorados_o_o_medio": 0.0,
        })
        item["qtd_referencias"] += 1
        item["media_acuracia_geral"] += row["acuracia_geral_percentual"]
        item["media_acuracia_entidade_referencia"] += row["acuracia_entidade_referencia_percentual"]
        item["media_acuracia_balanceada"] += row["acuracia_balanceada_percentual"]
        item["total_tokens_medio"] += row["total_tokens"]
        item["tokens_entidade_referencia_medio"] += row["tokens_entidade_referencia"]
        item["tokens_balanceados_medio"] += row["tokens_balanceados"]
        item["tokens_ignorados_o_o_medio"] += row["tokens_ignorados_o_o"]

    summary = []
    for item in summary_map.values():
        n = item["qtd_referencias"]
        item["media_acuracia_geral"] = item["media_acuracia_geral"] / n
        item["media_acuracia_entidade_referencia"] = item["media_acuracia_entidade_referencia"] / n
        item["media_acuracia_balanceada"] = item["media_acuracia_balanceada"] / n
        item["total_tokens_medio"] = item["total_tokens_medio"] / n
        item["tokens_entidade_referencia_medio"] = item["tokens_entidade_referencia_medio"] / n
        item["tokens_balanceados_medio"] = item["tokens_balanceados_medio"] / n
        item["tokens_ignorados_o_o_medio"] = item["tokens_ignorados_o_o_medio"] / n
        summary.append(item)
    summary.sort(key=lambda row: (row["horario"], row["candidato"], row["documento"]))

    account_rows = []
    for user, account in sorted(accounts.items(), key=lambda item: item[1]["candidato"]):
        account_rows.append({
            "usuario": user,
            "candidato": account["candidato"],
            "horario": account["horario"],
            "textos_planilha": account["textos_planilha"],
            "documentos_normalizados": "; ".join(account["documentos_planilha"]),
            "corrigido": "sim" if account["corrigido"] else "nao",
        })

    data = {
        "summary": summary,
        "details": sorted(details, key=lambda row: (row["documento"], row["candidato"], row["referencia"])),
        "skipped": skipped,
        "label_counts": label_counts,
        "accounts": account_rows,
    }
    (OUT_DIR / "accuracy_data.json").write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    for name, rows in [
        ("resumo.csv", summary),
        ("detalhes.csv", data["details"]),
        ("ignorados.csv", skipped),
        ("contagem_rotulos.csv", label_counts),
        ("contas.csv", account_rows),
    ]:
        path = OUT_DIR / name
        keys = sorted({key for row in rows for key in row.keys()})
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            writer.writerows(rows)

    print(json.dumps({
        "detalhes": len(details),
        "resumo": len(summary),
        "ignorados": len(skipped),
        "contas": len(account_rows),
        "saida": str(OUT_DIR / "accuracy_data.json"),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
